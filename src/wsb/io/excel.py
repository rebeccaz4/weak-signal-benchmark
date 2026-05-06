"""Excel output for evaluation summaries."""

from __future__ import annotations

from pathlib import Path

import pandas as pd


def save_summary(
    summary_df: pd.DataFrame, xlsx_path: Path, sheet_name: str
) -> None:
    """Save summary to Excel with proper sheet handling."""
    sheet_name = sheet_name[:31]  # Excel max sheet name length

    if xlsx_path.exists():
        from openpyxl import load_workbook

        book = load_workbook(xlsx_path)
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        with pd.ExcelWriter(
            xlsx_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            writer._book = book
            summary_df.to_excel(writer, sheet_name=sheet_name)
    else:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name=sheet_name)

    print(f"\nSaved summary to {xlsx_path} | sheet='{sheet_name}'")
