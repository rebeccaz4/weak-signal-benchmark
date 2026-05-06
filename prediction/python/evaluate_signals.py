#!/usr/bin/env python
"""Standalone LLM-as-a-Judge evaluation script.

Takes ground-truth signals and external-model signals as input, runs an LLM
judge N times with semantic matching, aggregates precision/recall/F1 with
mean/std, and saves results to Excel.

Example usage:
    python notebooks/backward/evaluate_signals.py \
      --ground-truth '["signal A", "signal B"]' \
      --external '["signal X", "signal Y"]' \
      --model-name dr_tulu \
      --topic "Reward Type - Process or Outcome" \
      --year-bucket "2020-2022" \
      --signal-type problem \
      --n-runs 2
"""
from __future__ import annotations

import argparse
import json
import os
import random
import re
import time
from pathlib import Path
from typing import Optional

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

# ---------------------------------------------------------------------------
# Prompts
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """\
You are a frontier researcher that is familiar with all specialty fields and have the highest IQ in the world.
You are also a careful, fair, and specialty-agnostic evaluator of research signals.
You must judge signals without favoring any domain; treat all fields as equally important.
Focus on the general qualities of each signal rather than your familiarity with its topic area.
"""

USER_PROMPT_TEMPLATE = """\
I will provide two signal sets:
1) Ground-truth signals (authoritative)
2) External-model signals to evaluate

Treat the ground-truth signals as correct weak signals.
Evaluate ONLY the external-model signals WITH RESPECT TO the ground-truth signals.

Compute and return:
- precision
- recall
- f1

Use semantic matching (not exact string matching). Match each external signal to at most one ground-truth signal.

Return ONLY valid JSON with this schema:
{{
"model": "{model_name}",
"precision": <float 0-1>,
"recall": <float 0-1>,
"f1": <float 0-1>,
"matched_pairs": [
  {{
    "external_signal": "<text>",
    "ground_truth_signal": "<text>"
  }}
],
"unmatched_external": ["<text>"],
"unmatched_ground_truth": ["<text>"]
}}

Ground-truth signals:
{ground_truth_block}

External-model signals:
{external_block}
"""

# ---------------------------------------------------------------------------
# Core functions
# ---------------------------------------------------------------------------


def format_signal_block(signals: list[str]) -> str:
    """Format signals as a numbered list."""
    return "\n".join(f"{i+1}. {s}" for i, s in enumerate(signals))


def build_eval_prompt(
    ground_truth: list[str],
    external: list[str],
    model_name: str = "external",
) -> str:
    """Build the user prompt for the LLM judge."""
    return USER_PROMPT_TEMPLATE.format(
        model_name=model_name,
        ground_truth_block=format_signal_block(ground_truth),
        external_block=format_signal_block(external),
    )


def _safe_json_loads(text: str) -> Optional[dict]:
    """Robust JSON parsing with regex fallback."""
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{[\s\S]*\}", text)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                return None
    return None


def run_judge_once(
    client: OpenAI,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 1.0,
    max_tokens: int = 10000,
    max_retries: int = 4,
    retry_backoff: float = 2.0,
) -> dict:
    """Single judge call with retry."""
    attempt = 0
    while True:
        attempt += 1
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=temperature,
                max_completion_tokens=max_tokens,
                response_format={"type": "json_object"},
            )

            if not resp.choices:
                raise ValueError("No choices returned from model.")

            text = (resp.choices[0].message.content or "").strip()
            if not text:
                finish = getattr(resp.choices[0], "finish_reason", None)
                refusal = getattr(
                    getattr(resp.choices[0], "message", None), "refusal", None
                )
                raise ValueError(
                    f"Empty response (finish_reason={finish}, refusal={refusal})."
                )

            payload = _safe_json_loads(text)
            if payload is None:
                raise ValueError("Could not parse JSON from model response.")
            return payload

        except Exception as exc:
            if attempt >= max_retries:
                raise RuntimeError(
                    f"LLM judge failed after {attempt} attempts: {exc}"
                ) from exc
            sleep_s = retry_backoff * attempt
            print(
                f"[warn] LLM error (attempt {attempt}): {exc}. "
                f"Retrying in {sleep_s:.1f}s..."
            )
            time.sleep(sleep_s)


def run_evaluation(
    ground_truth: list[str],
    external: list[str],
    n_runs: int = 10,
    model_name: str = "external",
    judge_model: str = "gpt-5",
    temperature: float = 1.0,
    max_tokens: int = 10000,
) -> pd.DataFrame:
    """Run N judge iterations and return a metrics DataFrame."""
    client = OpenAI()
    print("Building eval prompt...")
    user_prompt = build_eval_prompt(ground_truth, external, model_name)
    print("Eval prompt built.")
    import ipdb; ipdb.set_trace()

    print(f"Judge model: {judge_model} | temperature={temperature}")
    print(f"Ground-truth count: {len(ground_truth)}")
    print(f"External count:     {len(external)}")
    print(f"Running {n_runs} judge iterations...\n")

    raw_runs = []
    for i in range(1, n_runs + 1):
        print(f"  Run {i}/{n_runs}...", end=" ", flush=True)
        result = run_judge_once(
            client=client,
            model=judge_model,
            system_prompt=SYSTEM_PROMPT,
            user_prompt=user_prompt,
            temperature=temperature,
            max_tokens=max_tokens,
        )
        raw_runs.append(result)
        p = result.get("precision", "?")
        r = result.get("recall", "?")
        f = result.get("f1", "?")
        print(f"P={p}  R={r}  F1={f}")

    rows = []
    for run_idx, payload in enumerate(raw_runs, start=1):
        rows.append(
            {
                "run": run_idx,
                "model": payload.get("model", model_name),
                "precision": payload.get("precision"),
                "recall": payload.get("recall"),
                "f1": payload.get("f1"),
                "n_matched_pairs": len(payload.get("matched_pairs", []) or []),
                "n_unmatched_external": len(
                    payload.get("unmatched_external", []) or []
                ),
                "n_unmatched_ground_truth": len(
                    payload.get("unmatched_ground_truth", []) or []
                ),
            }
        )
    return pd.DataFrame(rows)


def save_summary(
    summary_df: pd.DataFrame, xlsx_path: Path, sheet_name: str
) -> None:
    """Save summary to Excel with proper sheet handling."""
    sheet_name = sheet_name[:31]  # Excel max sheet name length

    if xlsx_path.exists():
        from openpyxl import load_workbook

        book = load_workbook(xlsx_path)
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        with pd.ExcelWriter(
            xlsx_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            writer._book = book
            summary_df.to_excel(writer, sheet_name=sheet_name)
    else:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name=sheet_name)

    print(f"\nSaved summary to {xlsx_path} | sheet='{sheet_name}'")


def _load_signals(file_arg: str | None, inline_arg: str | None, label: str) -> list[str]:
    """Load signals from a JSON file path or inline JSON string."""
    if file_arg:
        path = Path(file_arg)
        if not path.exists():
            raise FileNotFoundError(f"{label} file not found: {path}")
        with open(path) as f:
            signals = json.load(f)
    elif inline_arg:
        signals = json.loads(inline_arg)
    else:
        raise ValueError(f"Must provide either --{label}-file or --{label}")

    if not isinstance(signals, list) or not all(isinstance(s, str) for s in signals):
        raise ValueError(f"{label} must be a JSON list of strings")
    return signals


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="LLM-as-a-Judge evaluation for weak-signal benchmarks.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    parser.add_argument(
        "--ground-truth-file",
        help="Path to JSON file with ground-truth signals",
    )
    parser.add_argument(
        "--ground-truth",
        help="Inline JSON list of ground-truth signals",
    )
    parser.add_argument(
        "--external-file",
        help="Path to JSON file with external-model signals",
    )
    parser.add_argument(
        "--external",
        help="Inline JSON list of external-model signals",
    )
    parser.add_argument(
        "--model-name",
        default="external",
        help="Label for the external model (default: external)",
    )
    parser.add_argument(
        "--topic",
        default="topic",
        help="Mainframe topic name (for sheet naming)",
    )
    parser.add_argument(
        "--year-bucket",
        default="",
        help='Year range, e.g. "2020-2022" (for sheet naming)',
    )
    parser.add_argument(
        "--signal-type",
        default="problem",
        choices=["problem", "solution"],
        help="Signal type (default: problem)",
    )
    parser.add_argument(
        "--judge-model",
        default=os.getenv("BWD_LLM_JUDGE_MODEL", "gpt-5"),
        help="Judge LLM model ID (default: gpt-5, env: BWD_LLM_JUDGE_MODEL)",
    )
    parser.add_argument(
        "--temperature",
        type=float,
        default=1.0,
        help="Judge temperature (default: 1.0)",
    )
    parser.add_argument(
        "--n-runs",
        type=int,
        default=10,
        help="Number of judge runs (default: 10)",
    )
    parser.add_argument(
        "--output-xlsx",
        help="Path to output Excel file (default: auto-generated)",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=27,
        help="Random seed (default: 27)",
    )

    args = parser.parse_args()
    random.seed(args.seed)

    # Load signals
    ground_truth = _load_signals(
        args.ground_truth_file, args.ground_truth, "ground-truth"
    )
    external = _load_signals(args.external_file, args.external, "external")

    # Run evaluation
    metrics_df = run_evaluation(
        ground_truth=ground_truth,
        external=external,
        n_runs=args.n_runs,
        model_name=args.model_name,
        judge_model=args.judge_model,
        temperature=args.temperature,
    )

    # Print per-run metrics
    print("\nPer-run metrics:")
    print(metrics_df.to_string(index=False))

    # Compute summary
    summary = (
        metrics_df[["precision", "recall", "f1"]]
        .agg(["mean", "std"])
        .T.round(4)
    )
    print("\nSummary (mean / std):")
    print(summary)

    # Determine output path
    if args.output_xlsx:
        xlsx_path = Path(args.output_xlsx)
    else:
        project_root = next(
            (
                p
                for p in Path.cwd().resolve().parents
                if (p / "README.md").exists()
            ),
            Path.cwd().resolve(),
        )
        eval_dir = project_root / "Evaluations"
        eval_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = eval_dir / f"llm_judge_summary_{args.model_name}.xlsx"

    # Build sheet name
    signal_label = args.signal_type.capitalize()
    sheet_name_raw = f"{signal_label} {args.year_bucket} {args.topic}"
    save_summary(summary, xlsx_path, sheet_name_raw)


if __name__ == "__main__":
    main()
